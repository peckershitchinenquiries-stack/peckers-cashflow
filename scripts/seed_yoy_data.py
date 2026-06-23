"""
Seed YoY weekly data from Excel files into Supabase tables.

Usage:
  cd C:/NextJs/peckers-cashflow
  python scripts/seed_yoy_data.py

Requires .env.local with:
  NEXT_PUBLIC_VM_SUPABASE_URL=https://jrmvvlkcwjaqruoyevyg.supabase.co
  VM_SUPABSE_SERVICE_ROLE_KEY=eyJ...  (service role key)
"""

import os, json
from pathlib import Path
import openpyxl
import requests
from dotenv import dotenv_values

# Config
env = dotenv_values(".env.local")
VM_SUPABASE_URL = env.get("NEXT_PUBLIC_VM_SUPABASE_URL", "").rstrip("/")
VM_SERVICE_KEY  = env.get("VM_SUPABSE_SERVICE_ROLE_KEY", "")

EXCEL_FILES = {
    "vm_yoy_both_stores": "yearly_data_both_stores_transformed.xlsx",
    "vm_yoy_hitchin":     "yearly_data_hitchin_transformed.xlsx",
    "vm_yoy_stevenage":   "yearly_data_stevenage_transformed.xlsx",
}

def load_excel(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for h, v in zip(headers, row):
            if h == "week_commencing":
                record[h] = str(v)[:10] if v else None
            else:
                record[h] = float(v) if v is not None else None
        if record.get("week_commencing"):
            rows.append(record)
    return rows

def upsert(table: str, rows: list[dict]):
    url = f"{VM_SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": VM_SERVICE_KEY,
        "Authorization": f"Bearer {VM_SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }
    batch_size = 500
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i+batch_size]
        r = requests.post(url, headers=headers, data=json.dumps(batch))
        if r.status_code not in (200, 201):
            raise RuntimeError(f"Upsert failed for {table}: {r.status_code} {r.text}")
        end = min(i + batch_size, len(rows))
        print(f"  {table}: upserted rows {i+1}-{end}")

if __name__ == "__main__":
    if not VM_SUPABASE_URL:
        raise SystemExit("ERROR: NEXT_PUBLIC_VM_SUPABASE_URL not found in .env.local")
    if not VM_SERVICE_KEY:
        raise SystemExit("ERROR: VM_SUPABSE_SERVICE_ROLE_KEY not found in .env.local")

    excel_dir = Path(__file__).parent.parent  # project root

    for table, filename in EXCEL_FILES.items():
        path = excel_dir / filename
        if not path.exists():
            print(f"  WARNING: Not found: {path} - skipping {table}")
            continue
        rows = load_excel(path)
        print(f"Seeding {table} - {len(rows)} rows from {filename}")
        upsert(table, rows)
        print(f"  Done")

    print("\nAll tables seeded.")
